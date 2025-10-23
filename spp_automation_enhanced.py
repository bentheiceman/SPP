"""
SPP Metric Automation - Enhanced Version v2.2 with User-Configurable Templates
Complete Multi-Tab Implementation with flexible template system.
        METRIC,eloper: Ben F. Benjamaa
Manager: Lauren B. Trapani
"""

import pandas as pd
import snowflake.connector
import os
import logging
import shutil
import json
from typing import List, Optional, Dict, Tuple
from datetime import datetime
import re
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from pathlib import Path

class SPPAutomationEnhanced:
    """Enhanced SPP Metric Automation with user-configurable templates and full multi-tab support."""
    
    def __init__(self, config_file: str = "config.ini", user_email: Optional[str] = None):
        self.config_file = config_file
        self.user_email = user_email
        self.connection: Optional[snowflake.connector.SnowflakeConnection] = None
        self.logger = self.setup_logging()
        self.template_config_file = "template_config.json"
        self.template_config = self.load_template_config()
        
    def setup_logging(self):
        """Set up logging configuration."""
        log_file = f"spp_automation_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler(log_file, encoding='utf-8'),
                logging.StreamHandler()
            ]
        )
        return logging.getLogger(__name__)
    
    def load_template_config(self) -> Dict:
        """Load template configuration from JSON file."""
        default_config = {
            "template_path": "",
            "use_template": False,
            "template_name": "SPP_Template.xlsm",
            "output_format": "xlsx",  # xlsx or xlsm
            "search_paths": [
                str(Path.home() / "OneDrive" / "Documents"),
                str(Path.home() / "Documents"),
                str(Path.cwd()),
                str(Path(__file__).parent)
            ]
        }
        
        try:
            if os.path.exists(self.template_config_file):
                with open(self.template_config_file, 'r') as f:
                    config = json.load(f)
                # Merge with default config to handle missing keys
                for key in default_config:
                    if key not in config:
                        config[key] = default_config[key]
                return config
            else:
                self.save_template_config(default_config)
                return default_config
        except Exception as e:
            self.logger.warning(f"Error loading template config: {e}. Using defaults.")
            return default_config
    
    def save_template_config(self, config: Dict) -> None:
        """Save template configuration to JSON file."""
        try:
            with open(self.template_config_file, 'w') as f:
                json.dump(config, f, indent=2)
            self.logger.info(f"Template configuration saved to {self.template_config_file}")
        except Exception as e:
            self.logger.error(f"Error saving template config: {e}")
    
    def update_template_config(self, template_path: str = "", use_template: bool = False, 
                             output_format: str = "xlsx") -> None:
        """Update template configuration."""
        self.template_config.update({
            "template_path": template_path,
            "use_template": use_template,
            "output_format": output_format,
            "last_updated": datetime.now().isoformat()
        })
        self.save_template_config(self.template_config)
        self.logger.info(f"Template config updated: Path={template_path}, Use={use_template}, Format={output_format}")
    
    def find_template_file(self) -> Optional[str]:
        """Find template file using configured path or search paths."""
        # First check if user has specified a custom path
        if self.template_config.get("template_path") and os.path.exists(self.template_config["template_path"]):
            self.logger.info(f"Using user-specified template: {self.template_config['template_path']}")
            return self.template_config["template_path"]
        
        # Search in predefined locations
        template_name = self.template_config.get("template_name", "SPP_Template.xlsm")
        search_paths = self.template_config.get("search_paths", [])
        
        for search_path in search_paths:
            template_path = os.path.join(search_path, template_name)
            if os.path.exists(template_path):
                self.logger.info(f"Found template at: {template_path}")
                return template_path
        
        self.logger.warning(f"Template file '{template_name}' not found in any search paths")
        return None
    
    def connect_to_snowflake(self) -> bool:
        """Connect to Snowflake using external browser authentication."""
        try:
            self.logger.info(f"Connecting to Snowflake with user: {self.user_email}")
            
            # Use EXACT same parameters that work in the test script
            self.connection = snowflake.connector.connect(
                user=self.user_email,
                account='HDSUPPLY-DATA',
                authenticator='externalbrowser',
                insecure_mode=True
            )
            
            # Set database context after connection
            cursor = self.connection.cursor()
            cursor.execute("USE DATABASE DM_SUPPLYCHAIN")
            cursor.execute("USE WAREHOUSE WH_SUPPLYCHAIN_ANALYST_XSMALL")
            cursor.execute("USE ROLE SUPPLYCHAIN_ANALYST")
            cursor.close()
            
            self.logger.info("Successfully connected to Snowflake and set context")
            return True
            
        except Exception as e:
            self.logger.error(f"Failed to connect to Snowflake: {e}")
            return False
    
    def get_query_0_summary_metrics(self, vendor_numbers: List[str], report_month: str, date_filter: str) -> str:
        """Generate Query 0 - Summary Metrics with immediate percentages including ASN success rate."""
        vendor_filter = "', '".join(vendor_numbers)
        
        return f"""
WITH Metric_Data AS (
    SELECT
        RPT_MONTH,
        VENDOR_NUMBER,
        VENDOR_NAME,
        Case 
            When Metric Like 'First_Receipt_FR_B1D' Then '1.Shipments_In_Full_1D'
            When Metric Like 'First_Receipt_FR_B28D' Then '2.Inbound_Fill_Rate_28D'
            When Metric Like 'Units_On_Time_Complete' Then '3.Units_On_Time_Complete'
        End As MetricType,
        TO_CHAR((SUM(METRIC_NUMERATOR)/SUM(METRIC_DENOMINATOR)) * 100, 'FM999.9') || '%' AS Metric_Percentage
    FROM VENDOR_PERFORMANCE.COMBINED_IPR_IB_VENDOR_PERFORMANCE
    WHERE 
        VENDOR_NUMBER IN ('{vendor_filter}')  -- Metric Data Vendor Filter
        AND RPT_MONTH like '{report_month}' -- Metric Data Month Filter
        AND METRIC IN ('First_Receipt_FR_B1D', 'First_Receipt_FR_B28D', 'Units_On_Time_Complete')
    GROUP BY RPT_MONTH, VENDOR_NUMBER, VENDOR_NAME, MetricType
),

ASN_Data AS (
    SELECT
        LTRIM(IH.LIFNR, 0) AS Vendor_Number,
        V.NAME1 AS Vendor_Name,
        TO_DATE(IH.ERDAT, 'YYYYMMDD') AS ERDAT_DATE,
        TO_CHAR((TO_DATE(IH.ERDAT, 'YYYYMMDD')), '"FY"YYYY-MON') AS RPT_MONTH,
        IH.VSTEL AS DC,
        ZUKRL AS PO_Number,
        LTRIM(IL.MATNR, 0) AS Material_Number,
        
        CASE 
            WHEN IH.VBELN LIKE '06%' AND IH.ERNAM IN ('BPAREMOTE', 'SCEBATCH', 'P2P_IDOC', 'P2PBATCH') THEN 'ASN'
            WHEN IH.VBELN LIKE '10%' THEN 'EGR'
            ELSE 'Manually Created'
        END AS Inbound_Type
    FROM EDP.STD_ECC.LIKP IH 
    
    INNER JOIN EDP.STD_ECC.LFA1 V
        ON IH.MANDT = V.MANDT
        AND IH.LIFNR = V.LIFNR
        
    INNER JOIN EDP.STD_ECC.LIPS IL
        ON IH.MANDT = IL.MANDT
        AND IH.VBELN = IL.VBELN
    WHERE IH.MANDT = '300'
        AND IH.LFART = 'ZEL'
        AND LTRIM(IH.LIFNR, 0) IN ('{vendor_filter}') -- ASN Supplier Filter
        AND IH.ERDAT LIKE '{date_filter}%' -- ASN Month Filter
),

ASN_Metric AS (
    SELECT
        RPT_MONTH,
        Vendor_Number,
        Vendor_Name,
        '4.ASN_Success_Rate' AS MetricType,
        TO_CHAR(COUNT(CASE WHEN Inbound_Type = 'ASN' THEN 1 END) * 100.0 / COUNT(*), 'FM999.9') || '%' AS Metric_Percentage
    FROM ASN_Data
    
    GROUP BY RPT_MONTH, Vendor_Number, Vendor_Name
)

SELECT * FROM Metric_Data
UNION ALL
SELECT * FROM ASN_Metric
ORDER BY VENDOR_NUMBER, MetricType
"""

    def get_query_1_basic_metrics(self, vendor_numbers: List[str], report_month: str) -> str:
        """Generate Query 1 - Enhanced Basic Metrics with receipt data and updated metric names."""
        vendor_filter = "', '".join(vendor_numbers)
        
        return f"""
WITH primary_metric AS (
    SELECT
        VENDOR_NUMBER,
        VENDOR_NAME,
        CONCAT(VENDOR_NUMBER,' - ',vendor_name) AS VENDOR,
        PO_NUMBER,
        USN,
        ITEM_DESCRIPTION,
        CONCAT(PO_NUMBER, ':', USN) AS Metric_Concatenate,
        CONCAT (USN, ' - ', ITEM_DESCRIPTION) as SKU,
        TO_DATE(DATE_ORIG_ORDERED) AS DATE_ORIG_ORDERED,
        TO_DATE(DATE_ORIG_PROMISED) AS DATE_ORIG_PROMISED,
        TO_DATE(DATE_FIRST_RECEIVED) AS DATE_FIRST_RECEIVED,
        WAREHOUSE_NUM,
        WAREHOUSE_NAME,
        CASE 
            WHEN Metric LIKE 'First_Receipt_FR_B1D' THEN 'Shipments_In_Full_1D'
            WHEN Metric LIKE 'First_Receipt_FR_B28D' THEN 'Inbound_Fill_Rate_28D'
            WHEN Metric LIKE 'Units_On_Time_Complete' THEN 'Units_On_Time_Complete'
        END AS MetricType,
        RPT_MONTH,
        FSCL_YR_PRD,
        METRIC_NUMERATOR,
        METRIC_DENOMINATOR,
        NETWORK
    FROM VENDOR_PERFORMANCE.COMBINED_IPR_IB_VENDOR_PERFORMANCE
    WHERE 
        VENDOR_NUMBER IN ('{vendor_filter}') -- Supplier Filter
        AND RPT_MONTH LIKE '{report_month}' -- Month Filter
        AND METRIC IN ('First_Receipt_FR_B1D', 'First_Receipt_FR_B28D', 'Units_On_Time_Complete')
),

hds_receipts AS (
    SELECT 
        CONCAT(e.ebeln, ':', LTRIM(e.MATNR, '0')) AS Metric_Concatenate,
        MAX(TRY_TO_DATE(TO_CHAR(e.budat), 'yyyymmdd')) AS receipt_date,
        a.MIC
    FROM EDP.STD_ECC.EKBE e
    LEFT JOIN IA_ATLAS.ATLAS a
        ON LTRIM(e.MATNR, '0') = LTRIM(a.MATERIAL, '0')
    WHERE e.bwart IN ('101', '102')
    GROUP BY e.ebeln, e.MATNR, a.MIC
),

hdp_receipts AS (
    SELECT
        CONCAT(PO_NUMBER, ':', USN) AS Metric_Concatenate,
        MAX(TO_DATE(DATE_RECEIVED)) AS receipt_date,
        MANUFACTURER_PART_NUMBER AS MIC
    FROM PRO_INVENTORY_ANALYTICS.REPORT_PURCHASE_ORDER_VISIBILITY_SHIPMENTS
    GROUP BY PO_NUMBER, USN, MANUFACTURER_PART_NUMBER
),

combined_receipts AS (
    SELECT
        Metric_Concatenate,
        MAX(receipt_date) AS receipt_date,
        MAX(MIC) AS MIC
    FROM (
        SELECT Metric_Concatenate, receipt_date, MIC FROM hds_receipts
        UNION ALL
        SELECT Metric_Concatenate, receipt_date, MIC FROM hdp_receipts
    ) all_receipts
    GROUP BY Metric_Concatenate
)

SELECT
    pm.RPT_MONTH as Report_Month,
    pm.NETWORK,
    pm.VENDOR,
    pm.WAREHOUSE_NUM,
    pm.WAREHOUSE_NAME,
    pm.PO_NUMBER,
    pm.SKU,
    cr.MIC as Vendor_Part_Number,
    pm.DATE_ORIG_ORDERED as Date_Ordered,
    pm.DATE_FIRST_RECEIVED,
    cr.receipt_date as Date_Last_Received,
    pm.MetricType,
    zeroifnull(pm.METRIC_NUMERATOR) as Metric_Units_Received,
    pm.METRIC_DENOMINATOR as Metric_Units_Ordered,
    case 
        when zeroifnull(pm.METRIC_NUMERATOR) < pm.METRIC_DENOMINATOR
        then 'Non-Compliant'
        else 'Compliant'
    end  as "Result"
    
FROM primary_metric pm
LEFT JOIN combined_receipts cr
    ON pm.Metric_Concatenate = cr.Metric_Concatenate
"""
    
    def get_query_2_asn_data(self, vendor_numbers: List[str], date_filter: str) -> str:
        """Generate Query 2 - ASN Data using LIKP/LIPS/LFA1 delivery tables."""
        vendor_filter = "', '".join(vendor_numbers)
        
        return f"""
SELECT
    CASE 
        WHEN IH.VBELN LIKE '06%' AND IH.ERNAM IN ('BPAREMOTE', 'SCEBATCH', 'P2P_IDOC', 'P2PBATCH') THEN 'ASN'
        WHEN IH.VBELN LIKE '10%' THEN 'EGR'
        ELSE 'Manually Created'
    END AS Inbound_Type,

    V.NAME1 AS Vendor_Name,
    LTRIM(IH.LIFNR, 0) AS Vendor_Number,
    TO_DATE(IH.ERDAT, 'YYYYMMDD') AS Create_Date,
    IH.VSTEL AS DC,
    ZUKRL AS PO_Number,
    M.MFRPN AS Vendor_Part_Number,
    LTRIM(IL.MATNR, 0) AS Material_Number,
    IL.ARKTX AS Material_Description,
    IL.ORMNG as Quantity_Ordered,
    IL.LFIMG AS Quantity_Received,
    IL.MEINS AS Unit_of_Measure,
    LTRIM(IH.VBELN, 0) AS Delivery_Number,
    LIFEX AS Supplier_Provided_ID,
    ZCARRIER_T AS Carrier_Name,
    BOLNR AS Provided_BOL,
    CASE 
        WHEN Inbound_Type = 'ASN' THEN 'Compliant'
        ELSE 'Non-Compliant'
    END AS "Result"

FROM EDP.STD_ECC.LIKP IH 

INNER JOIN EDP.STD_ECC.LIPS IL
    ON IH.MANDT = IL.MANDT
    AND IH.VBELN = IL.VBELN

INNER JOIN EDP.STD_ECC.LFA1 V
    ON IH.MANDT = V.MANDT
    AND IH.LIFNR = V.LIFNR

INNER JOIN EDP.STD_ECC.MARA M 
    ON IH.MANDT = M.MANDT
    AND IL.MATNR = M.MATNR

WHERE IH.MANDT = '300'
    AND IH.LFART = 'ZEL'
    AND LTRIM(IH.LIFNR, 0) IN ('{vendor_filter}') -- Ensure you're filtering the trimmed version
    AND IH.ERDAT LIKE '{date_filter}%' -- Filter for December 2024
"""
    
    def execute_query(self, query: str) -> pd.DataFrame:
        """Execute a query and return results as DataFrame."""
        try:
            if not self.connection:
                raise Exception("No active Snowflake connection")
                
            self.logger.info("Executing Snowflake query...")
            cursor = self.connection.cursor()
            
            # Execute query - now single statement since database context is set in connection
            cursor.execute(query)
            
            # Get column names
            columns = [desc[0] for desc in cursor.description] if cursor.description else []
            
            # Fetch data
            data = cursor.fetchall()
            
            # Create DataFrame
            df = pd.DataFrame(data, columns=columns)
            
            self.logger.info(f"Query executed successfully. Retrieved {len(df)} rows.")
            cursor.close()
            
            return df
            
        except Exception as e:
            self.logger.error(f"Error executing query: {e}")
            raise
    
    def create_output_directory(self) -> str:
        """Create output directory with timestamp."""
        output_dir = "Output"
        os.makedirs(output_dir, exist_ok=True)
        return output_dir
    
    def generate_filename(self, vendor_numbers: List[str], vendor_name: str, report_month: str) -> str:
        """Generate output filename based on vendor info and date."""
        # Clean vendor name for filename
        clean_vendor_name = re.sub(r'[<>:"/\\|?*]', '_', vendor_name.upper()) if vendor_name else "Unknown_Vendor"
        
        # Convert report month to readable format
        month_parts = report_month.split('-')
        if len(month_parts) == 2:
            year_part = month_parts[0].replace('FY', '')
            month_part = month_parts[1]
            readable_month = f"{month_part} {year_part}"
        else:
            readable_month = report_month
        
        # Determine file extension based on template
        ext = "xlsx"  # default
        if self.template_config.get("use_template", False):
            template_path = self.find_template_file()
            if template_path and template_path.endswith('.xlsm'):
                ext = "xlsm"
            elif template_path and template_path.endswith('.xlsx'):
                ext = "xlsx"
        
        # Create filename
        vendor_prefix = "_".join(vendor_numbers)
        filename = f"{vendor_prefix} - {clean_vendor_name} - {readable_month}.{ext}"
        
        return filename
    
    def get_vendor_name_from_data(self, df: pd.DataFrame) -> Optional[str]:
        """Extract vendor name from the DataFrame."""
        try:
            if 'VENDOR_NAME' in df.columns and not df.empty:
                vendor_names = df['VENDOR_NAME'].dropna().unique()
                if len(vendor_names) > 0:
                    return vendor_names[0]
            elif 'VENDOR' in df.columns and not df.empty:
                vendors = df['VENDOR'].dropna().unique()
                if len(vendors) > 0:
                    # Extract vendor name from "NUMBER - NAME" format
                    vendor_parts = vendors[0].split(' - ', 1)
                    return vendor_parts[1] if len(vendor_parts) > 1 else vendors[0]
        except Exception as e:
            self.logger.warning(f"Error extracting vendor name: {e}")
        
        return None
    
    def copy_template_file(self, output_path: str) -> bool:
        """Copy template file to output location."""
        try:
            template_path = self.find_template_file()
            if not template_path:
                self.logger.warning("No template file found, will create standard Excel file")
                return False
            
            shutil.copy2(template_path, output_path)
            self.logger.info(f"Template copied from {template_path} to {output_path}")
            return True
            
        except Exception as e:
            self.logger.error(f"Error copying template: {e}")
            return False
    
    def populate_template_tabs(self, output_path: str, data_dict: Dict[str, pd.DataFrame]) -> bool:
        """Populate template with data in different tabs."""
        try:
            workbook = openpyxl.load_workbook(output_path, keep_vba=True)
            
            # Define tab mapping - Summary first, then others
            tab_mapping = {
                'Summary_Metrics': 'Tab1_Summary_Metrics',
                'Basic_Metrics': 'Tab2_Basic_Metrics',
                'ASN_Data': 'Tab3_ASN_Data'
            }
            
            for data_key, sheet_name in tab_mapping.items():
                if data_key in data_dict and not data_dict[data_key].empty:
                    df = data_dict[data_key]
                    
                    # Create or get worksheet
                    if sheet_name in workbook.sheetnames:
                        worksheet = workbook[sheet_name]
                        # Clear existing data but keep formatting
                        for row in worksheet.iter_rows():
                            for cell in row:
                                if cell.row and cell.row > 1:  # Keep header row
                                    cell.value = None
                    else:
                        worksheet = workbook.create_sheet(sheet_name)
                    
                    # Write data starting from row 2 (assuming row 1 has headers)
                    start_row = 2 if sheet_name in workbook.sheetnames else 1
                    
                    # Write headers if new sheet
                    if start_row == 1:
                        for col, column_name in enumerate(df.columns, 1):
                            worksheet.cell(row=1, column=col, value=column_name)
                        start_row = 2
                    
                    # Write data
                    for row_idx, row_data in enumerate(df.itertuples(index=False), start_row):
                        for col_idx, value in enumerate(row_data, 1):
                            # Handle datetime objects
                            if pd.isna(value):
                                cell_value = None
                            elif isinstance(value, pd.Timestamp):
                                cell_value = value.strftime('%Y-%m-%d')
                            else:
                                cell_value = value
                            
                            worksheet.cell(row=row_idx, column=col_idx, value=cell_value)
                    
                    self.logger.info(f"Populated {sheet_name} with {len(df)} rows")
            
            # Remove default sheet if it exists and is empty
            if 'Sheet' in workbook.sheetnames and len(workbook.sheetnames) > 1:
                workbook.remove(workbook['Sheet'])
            
            workbook.save(output_path)
            workbook.close()
            self.logger.info(f"Template populated successfully: {output_path}")
            return True
            
        except Exception as e:
            self.logger.error(f"Error populating template: {e}")
            return False
    
    def create_standard_excel_file(self, output_path: str, data_dict: Dict[str, pd.DataFrame]) -> bool:
        """Create standard Excel file without template."""
        try:
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # Write each DataFrame to a separate sheet - Summary first, then others
                sheet_mapping = {
                    'Summary_Metrics': 'Tab1_Summary_Metrics',
                    'Basic_Metrics': 'Tab2_Basic_Metrics',
                    'ASN_Data': 'Tab3_ASN_Data'
                }
                
                for data_key, sheet_name in sheet_mapping.items():
                    if data_key in data_dict and not data_dict[data_key].empty:
                        df = data_dict[data_key]
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
                        self.logger.info(f"Created sheet {sheet_name} with {len(df)} rows")
            
            self.logger.info(f"Standard Excel file created successfully: {output_path}")
            return True
            
        except Exception as e:
            self.logger.error(f"Error creating standard Excel file: {e}")
            return False
    
    def run_full_automation(self, vendor_numbers: List[str], report_month: str, date_filter: str) -> Tuple[str, str]:
        """
        Run complete automation process.
        Returns: (output_file_path, status_message)
        """
        try:
            self.logger.info("=== Starting SPP Automation Process ===")
            self.logger.info(f"Vendors: {vendor_numbers}")
            self.logger.info(f"Report Month: {report_month}")
            self.logger.info(f"Date Filter: {date_filter}")
            
            # Connect to Snowflake
            if not self.connect_to_snowflake():
                return "", "Failed to connect to Snowflake"
            
            # Execute queries
            self.logger.info("Executing Summary Metrics query...")
            query0 = self.get_query_0_summary_metrics(vendor_numbers, report_month, date_filter)
            df_summary = self.execute_query(query0)
            
            self.logger.info("Executing Basic Metrics query...")
            query1 = self.get_query_1_basic_metrics(vendor_numbers, report_month)
            df_basic = self.execute_query(query1)
            
            self.logger.info("Executing ASN Data query...")
            query2 = self.get_query_2_asn_data(vendor_numbers, date_filter)
            df_asn = self.execute_query(query2)
            
            if df_summary.empty and df_basic.empty and df_asn.empty:
                return "", "No data found for the specified criteria"
            
            # Get vendor name
            vendor_name = (self.get_vendor_name_from_data(df_summary) or 
                          self.get_vendor_name_from_data(df_basic) or 
                          self.get_vendor_name_from_data(df_asn) or 
                          "Unknown_Vendor")
            
            # Create output directory and filename
            output_dir = self.create_output_directory()
            filename = self.generate_filename(vendor_numbers, vendor_name, report_month)
            output_path = os.path.join(output_dir, filename)
            
            # Prepare data dictionary with summary as first tab
            data_dict = {
                'Summary_Metrics': df_summary,
                'Basic_Metrics': df_basic,
                'ASN_Data': df_asn
            }
            
            # Create output file based on template configuration
            success = False
            creation_method = ""
            
            if self.template_config.get("use_template", False):
                # Try to use template
                template_path = self.find_template_file()
                if template_path and template_path.endswith('.xlsm'):
                    # Ensure output path has .xlsm extension for macro-enabled templates
                    if not output_path.endswith('.xlsm'):
                        output_path = output_path.replace('.xlsx', '.xlsm')
                    
                    if self.copy_template_file(output_path):
                        success = self.populate_template_tabs(output_path, data_dict)
                        creation_method = "macro-enabled template (.xlsm)"
                    else:
                        # Fallback to standard Excel if template copy failed
                        fallback_path = output_path.replace('.xlsm', '.xlsx')
                        success = self.create_standard_excel_file(fallback_path, data_dict)
                        output_path = fallback_path
                        creation_method = "standard Excel (template copy failed)"
                elif template_path and template_path.endswith('.xlsx'):
                    # Standard Excel template
                    if self.copy_template_file(output_path):
                        success = self.populate_template_tabs(output_path, data_dict)
                        creation_method = "Excel template (.xlsx)"
                    else:
                        # Fallback to standard Excel if template copy failed
                        success = self.create_standard_excel_file(output_path, data_dict)
                        creation_method = "standard Excel (template copy failed)"
                else:
                    # No template found, create standard Excel file
                    output_path = output_path.replace('.xlsm', '.xlsx')
                    success = self.create_standard_excel_file(output_path, data_dict)
                    creation_method = "standard Excel (template not found)"
            else:
                # Create standard Excel file
                output_path = output_path.replace('.xlsm', '.xlsx')
                success = self.create_standard_excel_file(output_path, data_dict)
                creation_method = "standard Excel"
            
            if success:
                status_msg = f"Successfully created {creation_method} file with {len(df_summary)} summary records, {len(df_basic)} basic metrics records and {len(df_asn)} ASN records"
                self.logger.info(f"=== Automation Complete ===")
                self.logger.info(f"Output file: {output_path}")
                return output_path, status_msg
            else:
                return "", "Failed to create output file"
                
        except Exception as e:
            error_msg = f"Automation failed: {str(e)}"
            self.logger.error(error_msg)
            return "", error_msg
        
        finally:
            if self.connection:
                self.connection.close()
                self.logger.info("Snowflake connection closed")
    
    def test_connection(self) -> Tuple[bool, str]:
        """Test Snowflake connection."""
        try:
            # Use EXACT same connection parameters as the working test script
            test_conn = snowflake.connector.connect(
                user=self.user_email,
                account='HDSUPPLY-DATA',
                authenticator='externalbrowser',
                insecure_mode=True
            )
            
            # Test with a simple query
            cursor = test_conn.cursor()
            cursor.execute("SELECT CURRENT_VERSION()")
            result = cursor.fetchone()
            cursor.close()
            test_conn.close()
            
            if result:
                return True, f"Connection successful. Snowflake version: {result[0]}"
            else:
                return False, "No result from test query"
                
        except Exception as e:
            return False, f"Connection test failed: {str(e)}"
    
    def test_query(self, vendor_numbers: List[str], report_month: str) -> Tuple[bool, str, int]:
        """Test queries without full execution."""
        try:
            if not self.connect_to_snowflake():
                return False, "Failed to connect to Snowflake", 0
            
            if not self.connection:
                return False, "Connection object is None", 0
            
            # Test basic metrics query
            query = self.get_query_1_basic_metrics(vendor_numbers, report_month)
            cursor = self.connection.cursor()
            cursor.execute(f"SELECT COUNT(*) FROM ({query})")
            result = cursor.fetchone()
            cursor.close()
            self.connection.close()
            
            if result:
                count = result[0]
                return True, f"Query test successful. Found {count} records.", count
            else:
                return False, "No result from test query", 0
            
        except Exception as e:
            return False, f"Query test failed: {str(e)}", 0

# Convenience function for backwards compatibility
def run_spp_automation(vendor_numbers: List[str], report_month: str, date_filter: str, 
                      user_email: str, template_path: str = "") -> Tuple[str, str]:
    """
    Convenience function to run SPP automation with template configuration.
    """
    automation = SPPAutomationEnhanced(user_email=user_email)
    
    # Configure template if provided
    if template_path:
        automation.update_template_config(
            template_path=template_path,
            use_template=True,
            output_format="xlsm"
        )
    
    return automation.run_full_automation(vendor_numbers, report_month, date_filter)

if __name__ == "__main__":
    # Example usage
    automation = SPPAutomationEnhanced()
    print("SPP Automation Enhanced - Ready for use")
    print(f"Template config: {automation.template_config_file}")
    print(f"Current template setting: {automation.template_config}")
