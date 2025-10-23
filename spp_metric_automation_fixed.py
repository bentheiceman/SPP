"""
SPP Metric Automation - Complete Multi-Tab Implementation
Fixes ASN data pulling and implements full multi-tab Excel functionality.

Developer: Ben F. Benjamaa
Manager: Lauren B. Trapani
"""

import pandas as pd
import snowflake.connector
import os
import logging
import shutil
from typing import List, Optional, Dict
from datetime import datetime
import re
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

class SPPMetricAutomationFixed:
    """Enhanced SPP Metric Automation with full multi-tab support and ASN integration."""
    
    def __init__(self, config_file: str, user_email: Optional[str] = None):
        self.config_file = config_file
        self.user_email = user_email
        self.connection = None
        self.logger = self.setup_logging()
        
    def setup_logging(self):
        """Set up logging configuration."""
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler('spp_automation_fixed.log', encoding='utf-8'),
                logging.StreamHandler()
            ]
        )
        return logging.getLogger(__name__)
    
    def connect_to_snowflake(self) -> bool:
        """Connect to Snowflake using external browser authentication."""
        try:
            self.logger.info(f"Connecting to Snowflake with user: {self.user_email}")
            
            self.connection = snowflake.connector.connect(
                account='HDSUPPLY-DATA',
                user=self.user_email,
                authenticator='externalbrowser',
                insecure_mode=True
            )
            
            self.logger.info("Successfully connected to Snowflake")
            return True
            
        except Exception as e:
            self.logger.error(f"Failed to connect to Snowflake: {e}")
            return False
    
    def get_query_1_basic_metrics(self, vendor_numbers: List[str], report_month: str) -> str:
        """Generate Query 1 - Basic Metrics with dynamic filters."""
        vendor_filter = "', '".join(vendor_numbers)
        
        return f"""
WITH primary_metric AS (
    SELECT
        VENDOR_NUMBER,
        VENDOR_NAME,
        CONCAT(VENDOR_NUMBER,' - ',vendor_name) AS VENDOR,
        PO_NUMBER,
        USN,
        ITEM_DESCRIPTION,
        CONCAT(PO_NUMBER, ':', USN) AS Metric_Concatenate,
        CONCAT (USN, ' - ', ITEM_DESCRIPTION) as SKU,
        TO_DATE(DATE_ORIG_ORDERED) AS DATE_ORIG_ORDERED,
        TO_DATE(DATE_ORIG_PROMISED) AS DATE_ORIG_PROMISED,
        TO_DATE(DATE_FIRST_RECEIVED) AS DATE_FIRST_RECEIVED,
        WAREHOUSE_NUM,
        WAREHOUSE_NAME,
        METRIC,
        RPT_MONTH,
        FSCL_YR_PRD,
        METRIC_NUMERATOR,
        METRIC_DENOMINATOR,
        NETWORK
    FROM DM_SUPPLYCHAIN.VENDOR_PERFORMANCE.COMBINED_IPR_IB_VENDOR_PERFORMANCE
    WHERE 
        VENDOR_NUMBER IN ('{vendor_filter}')
        AND RPT_MONTH = '{report_month}'
        AND METRIC IN ('First_Receipt_FR_B1D', 'First_Receipt_FR_B28D','Units_On_Time_Complete')
),

hds_receipts AS (
    SELECT 
        CONCAT(ebeln, ':', LTRIM(MATNR, '0')) AS Metric_Concatenate,
        MAX(TRY_TO_DATE(TO_CHAR(budat), 'yyyymmdd')) AS receipt_date
    FROM edp.std_ecc.ekbe
    WHERE bwart IN ('101', '102')
    GROUP BY ebeln, MATNR
),

hdp_receipts AS (
    SELECT
        CONCAT(PO_NUMBER, ':', USN) AS Metric_Concatenate,
        MAX(TO_DATE(DATE_RECEIVED)) AS receipt_date
    FROM DM_SUPPLYCHAIN.PRO_INVENTORY_ANALYTICS.REPORT_PURCHASE_ORDER_VISIBILITY_SHIPMENTS
    GROUP BY PO_NUMBER, USN
),

combined_receipts AS (
    SELECT
        Metric_Concatenate,
        MAX(receipt_date) AS receipt_date
    FROM (
        SELECT * FROM hds_receipts
        UNION ALL
        SELECT * FROM hdp_receipts
    ) all_receipts
    GROUP BY Metric_Concatenate
)

SELECT
    pm.RPT_MONTH as Report_Month,
    pm.NETWORK,
    pm.VENDOR,
    pm.WAREHOUSE_NUM,
    pm.WAREHOUSE_NAME,
    pm.PO_NUMBER,
    pm.SKU,
    pm.DATE_ORIG_ORDERED as Date_Ordered,
    pm.DATE_FIRST_RECEIVED,
    cr.receipt_date as Date_Last_Received,
    pm.METRIC,
    zeroifnull(pm.METRIC_NUMERATOR) as Metric_Units_Received,
    pm.METRIC_DENOMINATOR as Metric_Units_Ordered,
    case 
        when Metric_Units_Received < Metric_Units_Ordered
        then 'Non-Compliant'
        else 'Compliant'
    end  as "Result"
    
FROM primary_metric pm
LEFT JOIN combined_receipts cr
    ON pm.Metric_Concatenate = cr.Metric_Concatenate
ORDER BY pm.VENDOR, pm.PO_NUMBER, pm.SKU
"""

    def get_query_2_asn_data(self, vendor_numbers: List[str], date_filter: str) -> str:
        """Generate Query 2 - ASN Data with dynamic filters."""
        vendor_filter = "', '".join(vendor_numbers)
        
        return f"""
SELECT
    CASE 
        WHEN IH.VBELN LIKE '06%' AND IH.ERNAM IN ('BPAREMOTE', 'SCEBATCH', 'P2P_IDOC', 'P2PBATCH') THEN 'ASN'
        WHEN IH.VBELN LIKE '10%' THEN 'EGR'
        ELSE 'Manually Created'
    END AS Inbound_Type,

    V.NAME1 AS Vendor_Name,
    LTRIM(IH.LIFNR, 0) AS Vendor_Number,
    TO_DATE(IH.ERDAT, 'YYYYMMDD') AS Create_Date,
    IH.VSTEL AS DC,
    ZUKRL AS PO_Number,
    LTRIM(IL.MATNR, 0) AS Material_Number,
    IL.ARKTX AS Material_Description,
    IL.LFIMG AS Quantity,
    IL.MEINS AS Unit_of_Measure,
    LTRIM(IH.VBELN, 0) AS Delivery_Number,
    LIFEX AS Supplier_Provided_ID,
    ZCARRIER_T AS Carrier_Name,
    BOLNR AS Provided_BOL

FROM EDP.STD_ECC.LIKP IH 

INNER JOIN EDP.STD_ECC.LIPS IL
    ON IH.MANDT = IL.MANDT
    AND IH.VBELN = IL.VBELN

INNER JOIN EDP.STD_ECC.LFA1 V
    ON IH.MANDT = V.MANDT
    AND IH.LIFNR = V.LIFNR

WHERE IH.MANDT = '300'
    AND IH.LFART = 'ZEL'
    AND LTRIM(IH.LIFNR, 0) IN ('{vendor_filter}')
    AND IH.ERDAT LIKE '{date_filter}%'
ORDER BY IH.ERDAT DESC, V.NAME1, ZUKRL
"""

    def execute_query(self, query: str) -> pd.DataFrame:
        """Execute a query and return results as DataFrame."""
        if self.connection is None:
            raise Exception("No Snowflake connection available")
        
        try:
            self.logger.info("Executing query...")
            cursor = self.connection.cursor()
            cursor.execute(query)
            
            # Fetch results
            results = cursor.fetchall()
            columns = [desc[0] for desc in cursor.description]
            
            # Create DataFrame
            df = pd.DataFrame(results, columns=columns)
            cursor.close()
            
            self.logger.info(f"Query executed successfully. Retrieved {len(df)} rows.")
            return df
            
        except Exception as e:
            self.logger.error(f"Query execution failed: {e}")
            self.logger.error(f"Query: {query}")
            raise
    
    def create_output_directory(self) -> str:
        """Create output directory if it doesn't exist."""
        output_dir = os.path.join(os.getcwd(), "Output")
        os.makedirs(output_dir, exist_ok=True)
        return output_dir
    
    def generate_filename(self, vendor_numbers: List[str], vendor_name: str, report_month: str) -> str:
        """Generate filename based on vendor and month."""
        try:
            # Extract year and month from report_month (e.g., 'FY2025-APR')
            if 'FY' in report_month:
                parts = report_month.replace('FY', '').split('-')
                year = parts[0] if len(parts) > 0 else '2025'
                month = parts[1] if len(parts) > 1 else 'Unknown'
            else:
                year = '2025'
                month = report_month
            
            # Handle multiple vendors
            if len(vendor_numbers) == 1:
                vendor_part = vendor_numbers[0]
            else:
                vendor_part = '-'.join(vendor_numbers)
            
            # Clean vendor name for filename
            safe_vendor_name = re.sub(r'[^\w\s-]', '', vendor_name or 'Unknown_Vendor').strip()
            safe_vendor_name = re.sub(r'\s+', '_', safe_vendor_name)
            
            filename = f"{vendor_part} - {safe_vendor_name} - {month} {year}.xlsm"
            self.logger.info(f"Generated filename: {filename}")
            return filename
            
        except Exception as e:
            self.logger.error(f"Error generating filename: {e}")
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            return f"SPP_Metrics_{timestamp}.xlsm"
    
    def get_vendor_name_from_data(self, df: pd.DataFrame) -> Optional[str]:
        """Extract vendor name from data."""
        try:
            if df is not None and not df.empty:
                if 'VENDOR' in df.columns:
                    vendor_info = df['VENDOR'].iloc[0]
                    # Extract name part after ' - '
                    if ' - ' in vendor_info:
                        return vendor_info.split(' - ', 1)[1].strip()
                elif 'Vendor_Name' in df.columns:
                    return df['Vendor_Name'].iloc[0]
                elif 'VENDOR_NAME' in df.columns:
                    return df['VENDOR_NAME'].iloc[0]
            return None
        except Exception as e:
            self.logger.error(f"Error extracting vendor name: {e}")
            return None
    
    def copy_macro_template(self, output_path: str) -> bool:
        """Copy the macro template to output location."""
        try:
            # Try multiple possible template locations
            possible_template_paths = [
                r"C:\Users\1015723\OneDrive - HD Supply, Inc\Documents\Cole - Multi-Query information and template\Simplified Macro Template - SPP Monthly Details.xlsm",
                r"C:\Users\1015723\Documents\Cole - Multi-Query information and template\Simplified Macro Template - SPP Monthly Details.xlsm",
                r"C:\Users\1015723\Downloads\SPP\Simplified Macro Template - SPP Monthly Details.xlsm",
                r".\Simplified Macro Template - SPP Monthly Details.xlsm",
                # Add current directory as fallback
                os.path.join(os.path.dirname(__file__), "Simplified Macro Template - SPP Monthly Details.xlsm")
            ]
            
            template_path = None
            for path in possible_template_paths:
                if os.path.exists(path):
                    template_path = path
                    self.logger.info(f"Found template at: {template_path}")
                    break
            
            if not template_path:
                self.logger.warning("Template file not found in any of the expected locations:")
                for path in possible_template_paths:
                    self.logger.warning(f"  - {path}")
                return False
            
            # Ensure output directory exists
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            
            # Copy template to output location
            shutil.copy2(template_path, output_path)
            self.logger.info(f"Template copied to: {output_path}")
            return True
            
        except Exception as e:
            self.logger.error(f"Failed to copy template: {e}")
            return False

    def populate_template_tabs(self, output_path: str, data_dict: Dict[str, pd.DataFrame]) -> bool:
        """Populate data into the corresponding tabs of the macro template."""
        try:
            self.logger.info(f"Populating template tabs in: {output_path}")
            
            # Load the workbook with macros preserved
            from openpyxl import load_workbook
            wb = load_workbook(output_path, keep_vba=True)
            
            # Tab 1: METRIC DATA - populate starting at A1
            if 'METRIC DATA' in data_dict and not data_dict['METRIC DATA'].empty:
                self.logger.info("Populating METRIC DATA tab")
                
                # Check if sheet exists, create if not
                if 'METRIC DATA' not in wb.sheetnames:
                    wb.create_sheet('METRIC DATA')
                
                ws_metric = wb['METRIC DATA']
                
                # Clear existing data (preserve macros)
                for row in ws_metric.iter_rows():
                    for cell in row:
                        cell.value = None
                
                # Write headers
                df_metrics = data_dict['METRIC DATA']
                for col_idx, column in enumerate(df_metrics.columns, 1):
                    ws_metric.cell(row=1, column=col_idx, value=column)
                
                # Write data
                for row_idx, row_data in enumerate(df_metrics.values, 2):
                    for col_idx, value in enumerate(row_data, 1):
                        ws_metric.cell(row=row_idx, column=col_idx, value=value)
                
                self.logger.info(f"Populated METRIC DATA with {len(df_metrics)} rows")
            
            # Tab 2: ASN Data - populate starting at A1
            if 'ASN Data' in data_dict and not data_dict['ASN Data'].empty:
                self.logger.info("Populating ASN Data tab")
                
                # Check if sheet exists, create if not
                if 'ASN Data' not in wb.sheetnames:
                    wb.create_sheet('ASN Data')
                
                ws_asn = wb['ASN Data']
                
                # Clear existing data (preserve macros)
                for row in ws_asn.iter_rows():
                    for cell in row:
                        cell.value = None
                
                # Write headers
                df_asn = data_dict['ASN Data']
                for col_idx, column in enumerate(df_asn.columns, 1):
                    ws_asn.cell(row=1, column=col_idx, value=column)
                
                # Write data
                for row_idx, row_data in enumerate(df_asn.values, 2):
                    for col_idx, value in enumerate(row_data, 1):
                        ws_asn.cell(row=row_idx, column=col_idx, value=value)
                
                self.logger.info(f"Populated ASN Data with {len(df_asn)} rows")
            else:
                self.logger.warning("No ASN data to populate")
            
            # Save the workbook with macros preserved
            wb.save(output_path)
            self.logger.info("Template populated and saved successfully")
            
            return True
            
        except Exception as e:
            self.logger.error(f"Error populating template: {e}")
            return False

    def create_standard_excel_file(self, output_path: str, data_dict: Dict[str, pd.DataFrame]) -> bool:
        """Create a standard Excel file with multiple tabs when template is not available."""
        try:
            self.logger.info(f"Creating standard Excel file: {output_path}")
            
            # Create a new workbook
            from openpyxl import Workbook
            wb = Workbook()
            
            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            # Create METRIC DATA tab
            if 'METRIC DATA' in data_dict and not data_dict['METRIC DATA'].empty:
                self.logger.info("Writing METRIC DATA tab")
                ws_metric = wb.create_sheet('METRIC DATA')
                
                df_metrics = data_dict['METRIC DATA']
                
                # Write headers
                for col_idx, column in enumerate(df_metrics.columns, 1):
                    ws_metric.cell(row=1, column=col_idx, value=column)
                
                # Write data
                for row_idx, row_data in enumerate(df_metrics.values, 2):
                    for col_idx, value in enumerate(row_data, 1):
                        ws_metric.cell(row=row_idx, column=col_idx, value=value)
                
                self.logger.info(f"Created METRIC DATA tab with {len(df_metrics)} rows")
            
            # Create ASN Data tab
            if 'ASN Data' in data_dict and not data_dict['ASN Data'].empty:
                self.logger.info("Writing ASN Data tab")
                ws_asn = wb.create_sheet('ASN Data')
                
                df_asn = data_dict['ASN Data']
                
                # Write headers
                for col_idx, column in enumerate(df_asn.columns, 1):
                    ws_asn.cell(row=1, column=col_idx, value=column)
                
                # Write data
                for row_idx, row_data in enumerate(df_asn.values, 2):
                    for col_idx, value in enumerate(row_data, 1):
                        ws_asn.cell(row=row_idx, column=col_idx, value=value)
                
                self.logger.info(f"Created ASN Data tab with {len(df_asn)} rows")
            else:
                # Create empty ASN Data sheet with headers
                self.logger.warning("No ASN Data available - creating empty sheet")
                ws_asn = wb.create_sheet('ASN Data')
                asn_headers = ['Inbound_Type', 'Vendor_Name', 'Vendor_Number', 'Create_Date', 'DC', 
                              'PO_Number', 'Material_Number', 'Material_Description', 'Quantity', 
                              'Unit_of_Measure', 'Delivery_Number', 'Supplier_Provided_ID', 
                              'Carrier_Name', 'Provided_BOL']
                for col_idx, header in enumerate(asn_headers, 1):
                    ws_asn.cell(row=1, column=col_idx, value=header)
            
            # Ensure output directory exists
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            
            # Save the workbook
            wb.save(output_path)
            self.logger.info("Standard Excel file created successfully")
            
            return True
            
        except Exception as e:
            self.logger.error(f"Error creating standard Excel file: {e}")
            return False
    
    def run_full_automation(self, vendor_numbers: List[str], report_month: str, date_filter: str) -> str:
        """Run the complete automation process with enhanced error handling and logging."""
        try:
            self.logger.info(f"=== Starting Full SPP Automation ===")
            self.logger.info(f"Vendors: {vendor_numbers}")
            self.logger.info(f"Report Month: {report_month}")
            self.logger.info(f"Date Filter: {date_filter}")
            
            # Step 1: Connect to Snowflake
            if not self.connect_to_snowflake():
                raise Exception("Failed to connect to Snowflake")
            
            # Step 2: Execute Query 1 - Basic Metrics
            self.logger.info("=== Executing Query 1 - Basic Metrics ===")
            query1 = self.get_query_1_basic_metrics(vendor_numbers, report_month)
            self.logger.debug(f"Query 1: {query1}")
            df_metrics = self.execute_query(query1)
            
            if df_metrics.empty:
                self.logger.warning(f"No metric data found for vendors {vendor_numbers} and month {report_month}")
            else:
                self.logger.info(f"Retrieved {len(df_metrics)} metric records")
            
            # Step 3: Execute Query 2 - ASN Data
            self.logger.info("=== Executing Query 2 - ASN Data ===")
            query2 = self.get_query_2_asn_data(vendor_numbers, date_filter)
            self.logger.debug(f"Query 2: {query2}")
            df_asn = self.execute_query(query2)
            
            if df_asn.empty:
                self.logger.warning(f"No ASN data found for vendors {vendor_numbers} and date filter {date_filter}")
            else:
                self.logger.info(f"Retrieved {len(df_asn)} ASN records")
            
            # Step 4: Get vendor name and create filename
            vendor_name = self.get_vendor_name_from_data(df_metrics) or self.get_vendor_name_from_data(df_asn)
            if not vendor_name:
                vendor_name = "Unknown_Vendor"
                self.logger.warning("Could not determine vendor name from data")
            
            # Step 5: Create output directory and filename
            output_dir = self.create_output_directory()
            filename = self.generate_filename(vendor_numbers, vendor_name, report_month)
            output_path = os.path.join(output_dir, filename)
            
            # Step 6: Prepare data dictionary for Excel population
            data_dict = {
                'METRIC DATA': df_metrics,
                'ASN Data': df_asn
            }
            
            # Step 7: Try to use macro template, fall back to standard Excel if needed
            copy_success = self.copy_macro_template(output_path)
            
            if copy_success:
                # Template available - use template-based approach
                populate_success = self.populate_template_tabs(output_path, data_dict)
                if not populate_success:
                    self.logger.warning("Failed to populate template, falling back to standard Excel")
                    populate_success = self.create_standard_excel_file(output_path, data_dict)
            else:
                # Template not available - use standard Excel approach
                self.logger.warning("Macro template not available, creating standard Excel file")
                populate_success = self.create_standard_excel_file(output_path, data_dict)
            
            # Step 8: Close Snowflake connection
            if self.connection:
                self.connection.close()
                self.logger.info("Snowflake connection closed")
            
            if populate_success:
                self.logger.info(f"=== Automation completed successfully ===")
                self.logger.info(f"Output file: {output_path}")
                if copy_success:
                    self.logger.info("NEXT STEPS:")
                    self.logger.info("1. Open the Excel file")
                    self.logger.info("2. Press Ctrl+Shift+M to run the macro")
                    self.logger.info("3. The macro will refresh pivots and populate percentages")
                    return f"Automation completed successfully! Macro-enabled file created: {output_path}. Use Ctrl+Shift+M to run macros."
                else:
                    self.logger.info("Standard Excel file created (macro template not available)")
                    return f"Automation completed successfully! Standard Excel file created: {output_path}. Note: Macro template was not available."
            else:
                raise Exception("Failed to create Excel file")
                
        except Exception as e:
            error_msg = f"Automation failed: {str(e)}"
            self.logger.error(error_msg)
            if self.connection:
                self.connection.close()
            return error_msg

    def test_asn_query_standalone(self, vendor_numbers: List[str], date_filter: str) -> pd.DataFrame:
        """Test ASN query independently for debugging."""
        try:
            self.logger.info(f"=== Testing ASN Query Standalone ===")
            self.logger.info(f"Vendors: {vendor_numbers}, Date Filter: {date_filter}")
            
            if not self.connect_to_snowflake():
                raise Exception("Failed to connect to Snowflake")
            
            query = self.get_query_2_asn_data(vendor_numbers, date_filter)
            self.logger.info("ASN Query:")
            self.logger.info(query)
            
            df = self.execute_query(query)
            self.logger.info(f"ASN Query Results: {len(df)} rows")
            
            if not df.empty:
                self.logger.info("Sample ASN data:")
                self.logger.info(df.head().to_string())
            
            if self.connection:
                self.connection.close()
                
            return df
            
        except Exception as e:
            self.logger.error(f"ASN query test failed: {e}")
            if self.connection:
                self.connection.close()
            raise
