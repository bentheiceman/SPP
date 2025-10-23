#!/usr/bin/env python3
"""
Simple test for macro template functionality
"""

import os
import json
import openpyxl

def test_macro_template_basic():
    """Basic test for macro template support."""
    
    print("Testing Macro Template Support")
    print("=" * 40)
    
    # Check if macro template exists
    macro_template_path = r"C:\Users\1015723\OneDrive - HD Supply, Inc\Documents\Cole - Multi-Query information and template\Simplified Macro Template - SPP Monthly Details.xlsm"
    
    if not os.path.exists(macro_template_path):
        print(f"❌ Template not found: {macro_template_path}")
        return False
    
    print(f"✅ Found template: {os.path.basename(macro_template_path)}")
    
    # Test openpyxl macro support
    try:
        print("Testing openpyxl macro loading...")
        workbook = openpyxl.load_workbook(macro_template_path, keep_vba=True)
        print("✅ Successfully loaded macro-enabled workbook")
        
        # Check sheets
        sheets = workbook.sheetnames
        print(f"✅ Found {len(sheets)} sheets: {sheets}")
        
        # Check for VBA
        if hasattr(workbook, 'vba_archive') and workbook.vba_archive:
            print("✅ VBA macros detected and preserved")
        else:
            print("⚠️ No VBA archive detected")
        
        workbook.close()
        
    except Exception as e:
        print(f"❌ Error loading macro file: {e}")
        return False
    
    # Test copying and saving as macro-enabled
    try:
        print("Testing macro file copy and save...")
        test_output = "test_macro_output.xlsm"
        
        # Load original
        workbook = openpyxl.load_workbook(macro_template_path, keep_vba=True)
        
        # Save copy
        workbook.save(test_output)
        workbook.close()
        
        # Verify copy
        if os.path.exists(test_output):
            print(f"✅ Successfully created macro copy: {test_output}")
            
            # Test loading the copy
            copy_workbook = openpyxl.load_workbook(test_output, keep_vba=True)
            if hasattr(copy_workbook, 'vba_archive') and copy_workbook.vba_archive:
                print("✅ Macros preserved in copy")
            else:
                print("⚠️ Macros may not be preserved in copy")
            copy_workbook.close()
            
            # Clean up
            os.remove(test_output)
            print("✅ Cleanup completed")
            
        else:
            print("❌ Failed to create macro copy")
            return False
            
    except Exception as e:
        print(f"❌ Error in copy/save test: {e}")
        return False
    
    print("\n🎉 All basic macro tests passed!")
    return True

if __name__ == "__main__":
    success = test_macro_template_basic()
    if success:
        print("\n✅ Your system supports macro-enabled templates!")
    else:
        print("\n❌ Macro template support issues detected")
